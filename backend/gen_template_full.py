"""Generate Equipment Assignment Import Template (full, with reference columns)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# Colors and styles
HEADER_FILL = PatternFill("solid", fgColor="1565C0")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
EXAMPLE_FILL = PatternFill("solid", fgColor="FFF9C4")
EXAMPLE_FONT = Font(italic=True, color="999999")
INSTR_TITLE_FONT = Font(bold=True, size=14, color="1565C0")
INSTR_SECTION_FONT = Font(bold=True, size=11)
CENTER = Alignment(horizontal="center", vertical="center")

# 1. IMPORT sheet
ws = wb.active
ws.title = "Import"
ws.sheet_properties.tabColor = "4CAF50"

columns = [
    ("Equipment ID (required)", 18),
    ("Equipment Name", 28),
    ("Category", 20),
    ("Employee ID (required)", 16),
    ("Employee Name", 28),
    ("Location (required)", 25),
    ("Customer", 25),
    ("Checkout Date (required, YYYY-MM-DD)", 28),
    ("Expected Return Date", 22),
    ("Notes", 35),
]

for col_idx, (header, width) in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[1].height = 28

# Example rows
examples = [
    ["EQ-VA-001", "Vibration Analyser", "Vibration", "EMP001", "John Smith", "Middelburg Lab", "Valterra - CP", "2026-01-15", "2026-06-30", "Example - delete before importing"],
    ["EQ-TC-002", "Flir Camera", "Thermal Camera", "EMP002", "Trevor Komane", "Secunda Site", "Sasol", "2026-02-01", "2026-06-30", "Example - delete before importing"],
]
for row_idx, data in enumerate(examples, 2):
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = EXAMPLE_FONT
        cell.fill = EXAMPLE_FILL
ws.freeze_panes = "A2"

# 2. INSTRUCTIONS sheet
instr_ws = wb.create_sheet("Instructions")
instr_ws.sheet_properties.tabColor = "FFAB40"
instr_ws.column_dimensions["A"].width = 110

lines = [
    ("EQUIPMENT ASSIGNMENT IMPORT TEMPLATE", INSTR_TITLE_FONT),
    ("", None),
    ("MANDATORY FIELDS:", INSTR_SECTION_FONT),
    ("- Equipment ID (must match an existing Equipment ID in the system)", None),
    ("- Employee ID (must match an existing Employee ID in the system)", None),
    ("- Location (must match an existing Location name exactly)", None),
    ("- Checkout Date (YYYY-MM-DD)", None),
    ("", None),
    ("OPTIONAL FIELDS:", INSTR_SECTION_FONT),
    ("- Equipment Name, Category, Employee Name: for your reference only (ignored by import)", None),
    ("- Customer, Expected Return Date, Notes: optional for planning and info", None),
    ("", None),
    ("HOW TO USE:", INSTR_SECTION_FONT),
    ("1. Delete the 2 yellow example rows on the Import sheet", None),
    ("2. Fill in your data starting from row 2 (row 1 is the header — do not change it)", None),
    ("3. Use exact Equipment IDs, Employee IDs, and Location names as they appear in the system", None),
    ("4. Dates must be entered as YYYY-MM-DD (e.g. 2026-03-15)", None),
    ("5. Save the file as .xlsx (keep the format)", None),
    ("6. Upload via the Equipment Store app", None),
    ("", None),
    ("WHAT HAPPENS ON IMPORT:", INSTR_SECTION_FONT),
    ("- Equipment status is set to 'Checked Out'", None),
    ("- Employee is assigned as current holder", None),
    ("- Location is updated on the equipment record", None),
    ("- A checkout movement record is created with your specified date", None),
    ("- If an Equipment ID appears multiple times, only the last occurrence is used", None),
]
for row_idx, (text, font) in enumerate(lines, 1):
    cell = instr_ws.cell(row=row_idx, column=1, value=text)
    if font:
        cell.font = font

# Save
output = os.path.join(
    "C:\\", "Users", "nadhi",
    "OneDrive - Wearcheck Reliability Solutions",
    "Desktop",
    "Equipment_Assignment_Import_Template_FULL.xlsx"
)
wb.save(output)
print(f"\nTemplate saved to:\n  {output}")
